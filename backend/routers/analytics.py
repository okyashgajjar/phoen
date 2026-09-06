"""
Admin Reporting — mockup Screen 15, spec A7.

Sales performance with the four filters the spec names (Period, Sales Team /
Rep, Approval Status, Product / Category) and PDF / XLS export.

Everything is computed from the documents themselves. Where a figure cannot be
derived from the data, the response says so rather than substituting a
plausible-looking constant -- the dashboard endpoint next door still returns a
hardcoded `win_velocity_days: 11.4`, which is exactly the habit this avoids.
"""

import io
import os
import sys
from collections import Counter, defaultdict
from datetime import timezone, datetime, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse

ROOT_DIR = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if ROOT_DIR not in sys.path:
    sys.path.insert(0, ROOT_DIR)

from database.config import SessionLocal
from database.models import SalesDocument, DocumentLine, CatalogItem, Variant, Category, AuditLog, Customer
from models.base import db
from models.sales import normalize_status
from models.users import RoleEnum
from dependencies import RoleChecker

router = APIRouter()

PERIODS = {"today": 1, "week": 7, "month": 30, "quarter": 91, "year": 365}

APPROVAL_BUCKETS = {
    "pending": {"PENDING_APPROVAL"},
    "approved": {"APPROVED", "READY", "CONFIRMED", "WON", "DISPATCHED", "PAID"},
    "rejected": {"REJECTED", "EXPIRED"},
}


def _f(v, d=0.0):
    try:
        return float(v)
    except (TypeError, ValueError):
        return d


def _parse_date(value: Optional[str]) -> Optional[datetime]:
    if not value:
        return None
    try:
        return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return None


def _window(period: str, date_from: Optional[str], date_to: Optional[str]):
    start = _parse_date(date_from)
    end = _parse_date(date_to)
    if start or end:
        return start, end
    days = PERIODS.get((period or "").lower())
    if days:
        return datetime.now(timezone.utc) - timedelta(days=days), None
    return None, None


def _collect(session, period, date_from, date_to, rep, approval_status, category_id, product_id):
    """Apply the four filters and return the matching quotations plus their lines."""
    start, end = _window(period, date_from, date_to)

    q = session.query(SalesDocument).filter(SalesDocument.document_type == "QUOTATION")
    if start:
        q = q.filter(SalesDocument.document_date >= start)
    if end:
        q = q.filter(SalesDocument.document_date <= end)
    if rep:
        q = q.filter(SalesDocument.created_by == rep)

    docs = q.all()

    if approval_status and isinstance(approval_status, str) and approval_status.strip():
        allowed = APPROVAL_BUCKETS.get(approval_status.strip().lower())
        if allowed:
            docs = [d for d in docs if normalize_status(d.status) in allowed]

    doc_ids = [d.id for d in docs]
    lines = []
    if doc_ids:
        lines = (
            session.query(DocumentLine)
            .filter(DocumentLine.document_id.in_(doc_ids))
            .all()
        )

    # Product / category filter narrows to quotations containing a match.
    if category_id or product_id:
        items = {i.id: i for i in session.query(CatalogItem).all()}
        keep = set()
        for line in lines:
            item = items.get(line.catalog_item_id)
            if product_id and line.catalog_item_id == product_id:
                keep.add(line.document_id)
            elif category_id and item is not None and item.category_id == category_id:
                keep.add(line.document_id)
        docs = [d for d in docs if d.id in keep]
        doc_ids = set(d.id for d in docs)
        lines = [l for l in lines if l.document_id in doc_ids]

    return docs, lines


def _avg_approval_hours(session, docs) -> Optional[float]:
    """
    Mean hours from submission to an approval decision, taken from the audit
    ledger. Returns None when no approval events are on record, rather than
    inventing a number.
    """
    ids = [d.id for d in docs]
    if not ids:
        return None

    events = (
        session.query(AuditLog)
        .filter(AuditLog.entity_type.in_(["QUOTATION", "quotation"]))
        .filter(AuditLog.entity_id.in_(ids))
        .all()
    )
    by_doc = defaultdict(list)
    for e in events:
        if e.timestamp:
            by_doc[e.entity_id].append((e.timestamp, (e.action or "").upper()))

    spans = []
    for doc in docs:
        history = sorted(by_doc.get(doc.id, []))
        submitted = next((t for t, a in history if "SUBMIT" in a), None)
        decided = next((t for t, a in history if "APPROV" in a or "REJECT" in a), None)
        if submitted and decided and decided >= submitted:
            spans.append((decided - submitted).total_seconds() / 3600.0)

    if not spans:
        return None
    return round(sum(spans) / len(spans), 1)


@router.get("/analytics")
def sales_analytics(
    period: str = "month",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rep: Optional[str] = None,
    approval_status: Optional[str] = None,
    category_id: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(
        RoleChecker([RoleEnum.admin, RoleEnum.manager, RoleEnum.finance])
    ),
):
    session = SessionLocal()
    try:
        docs, lines = _collect(
            session, period, date_from, date_to, rep, approval_status, category_id, product_id
        )

        items = {i.id: i for i in session.query(CatalogItem).all()}
        categories = {c.id: c.name for c in session.query(Category).all()}

        statuses = Counter(normalize_status(d.status) for d in docs)
        total_value = sum(_f(d.grand_total) for d in docs)
        won = [d for d in docs if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID")]
        decided = [
            d for d in docs
            if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID", "REJECTED", "EXPIRED")
        ]

        # Per-product performance
        per_product = defaultdict(lambda: {"qty": 0, "revenue": 0.0, "discount_sum": 0.0, "lines": 0})
        for line in lines:
            item = items.get(line.catalog_item_id)
            key = line.catalog_item_id or "UNLINKED"
            entry = per_product[key]
            entry["qty"] += int(line.quantity or 0)
            entry["revenue"] += _f(line.line_total)
            entry["discount_sum"] += _f(line.discount_percent)
            entry["lines"] += 1
            entry["name"] = item.name if item is not None else "Unlinked line"
            entry["category"] = categories.get(item.category_id) if item is not None else None

        products = [
            {
                "product_id": pid,
                "name": v.get("name"),
                "category": v.get("category"),
                "units": v["qty"],
                "revenue": round(v["revenue"], 2),
                "avg_discount": round(v["discount_sum"] / v["lines"], 2) if v["lines"] else 0.0,
                "line_count": v["lines"],
            }
            for pid, v in per_product.items()
        ]
        # Lines with no catalog link are not a product. Ranking them alongside
        # real products puts a phantom "Unlinked line" at the top of every
        # chart, so they are excluded and reported separately as a data-quality
        # figure instead.
        unlinked = per_product.get("UNLINKED", {})
        ranked = [p for p in products if p["product_id"] != "UNLINKED"]
        best_selling = sorted(ranked, key=lambda p: p["revenue"], reverse=True)[:10]
        most_discounted = sorted(ranked, key=lambda p: p["avg_discount"], reverse=True)[:10]

        # Per-rep performance
        per_rep = defaultdict(lambda: {"quotes": 0, "value": 0.0, "won": 0})
        for d in docs:
            who = d.created_by or "Unassigned"
            per_rep[who]["quotes"] += 1
            per_rep[who]["value"] += _f(d.grand_total)
            if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID"):
                per_rep[who]["won"] += 1
        reps = sorted(
            [
                {
                    "rep": k,
                    "quotes": v["quotes"],
                    "value": round(v["value"], 2),
                    "won": v["won"],
                    "win_rate": round(v["won"] / v["quotes"] * 100, 1) if v["quotes"] else 0.0,
                }
                for k, v in per_rep.items()
            ],
            key=lambda r: r["value"],
            reverse=True,
        )

        # Customer & Tier analytics
        customers = {c.id: c for c in session.query(Customer).all()}
        tier_stats = defaultdict(lambda: {"quotes": 0, "value": 0.0, "won_val": 0.0, "won_count": 0, "discount_sum": 0.0, "line_count": 0})
        for d in docs:
            c = customers.get(d.customer_id)
            tier = c.tier if c and c.tier else "Standard"
            tier_stats[tier]["quotes"] += 1
            val = _f(d.grand_total)
            tier_stats[tier]["value"] += val
            if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID"):
                tier_stats[tier]["won_val"] += val
                tier_stats[tier]["won_count"] += 1

        for l in lines:
            d_owner = next((d for d in docs if d.id == l.document_id), None)
            if d_owner:
                c = customers.get(d_owner.customer_id)
                tier = c.tier if c and c.tier else "Standard"
                tier_stats[tier]["discount_sum"] += _f(l.discount_percent)
                tier_stats[tier]["line_count"] += 1

        tier_breakdown = [
            {
                "tier": t,
                "quotes": v["quotes"],
                "value": round(v["value"], 2),
                "won_value": round(v["won_val"], 2),
                "won_count": v["won_count"],
                "win_rate": round(v["won_count"] / v["quotes"] * 100, 1) if v["quotes"] else 0.0,
                "avg_discount": round(v["discount_sum"] / v["line_count"], 1) if v["line_count"] else 0.0,
            }
            for t, v in sorted(tier_stats.items(), key=lambda kv: -kv[1]["value"])
        ]

        # Top customer accounts
        cust_stats = defaultdict(lambda: {"name": "", "tier": "", "quotes": 0, "value": 0.0, "won_val": 0.0})
        for d in docs:
            c = customers.get(d.customer_id)
            cid = d.customer_id
            cust_stats[cid]["name"] = c.company_name if c else "Unknown Customer"
            cust_stats[cid]["tier"] = c.tier if c and c.tier else "Standard"
            cust_stats[cid]["quotes"] += 1
            val = _f(d.grand_total)
            cust_stats[cid]["value"] += val
            if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID"):
                cust_stats[cid]["won_val"] += val

        top_customers = sorted(
            [
                {
                    "customer_id": cid,
                    "name": v["name"],
                    "tier": v["tier"],
                    "quotes": v["quotes"],
                    "value": round(v["value"], 2),
                    "won_value": round(v["won_val"], 2),
                }
                for cid, v in cust_stats.items()
            ],
            key=lambda x: -x["value"]
        )[:10]

        # Funnel breakdown
        funnel_map = {
            "DRAFT": "Draft Proposals",
            "PENDING_APPROVAL": "Pending Review",
            "APPROVED": "Approved & Ready",
            "READY": "Ready for Customer",
            "NEGOTIATION": "Active Negotiation",
            "CONFIRMED": "Confirmed Orders",
            "WON": "Won & Signed",
            "REJECTED": "Rejected",
            "EXPIRED": "Expired",
        }
        funnel_totals = defaultdict(lambda: {"count": 0, "value": 0.0})
        for d in docs:
            st = normalize_status(d.status)
            stage_name = funnel_map.get(st, st)
            funnel_totals[stage_name]["count"] += 1
            funnel_totals[stage_name]["value"] += _f(d.grand_total)

        funnel_stages = [
            {
                "stage": k,
                "count": v["count"],
                "value": round(v["value"], 2),
                "share_pct": round(v["value"] / total_value * 100, 1) if total_value > 0 else 0.0,
            }
            for k, v in sorted(funnel_totals.items(), key=lambda kv: -kv[1]["value"])
        ]

        # Discount Leakage metrics
        total_discount_amount = sum(_f(l.discount_amount) for l in lines)
        high_disc_lines = [l for l in lines if _f(l.discount_percent) > 15.0]
        critical_disc_lines = [l for l in lines if _f(l.discount_percent) > 20.0]

        # Monthly timeline trend
        by_month = defaultdict(lambda: {"quotes": 0, "value": 0.0, "won_val": 0.0})
        for d in docs:
            m = d.document_date.strftime("%Y-%m") if d.document_date else "Recent"
            by_month[m]["quotes"] += 1
            val = _f(d.grand_total)
            by_month[m]["value"] += val
            if normalize_status(d.status) in ("WON", "CONFIRMED", "PAID"):
                by_month[m]["won_val"] += val

        monthly_trend = [
            {
                "period": m,
                "quotes": v["quotes"],
                "value": round(v["value"], 2),
                "won_value": round(v["won_val"], 2),
            }
            for m, v in sorted(by_month.items())
        ]

        avg_hours = _avg_approval_hours(session, docs)
        discounts = [_f(l.discount_percent) for l in lines if l.discount_percent is not None]

        return {
            "filters": {
                "period": period,
                "date_from": date_from,
                "date_to": date_to,
                "rep": rep,
                "approval_status": approval_status,
                "category_id": category_id,
                "product_id": product_id,
            },
            "kpis": {
                "quotes_created": len(docs),
                "total_value": round(total_value, 2),
                "average_deal_size": round(total_value / len(docs), 2) if docs else 0.0,
                "won_count": len(won),
                "won_value": round(sum(_f(d.grand_total) for d in won), 2),
                "win_rate": round(len(won) / len(decided) * 100, 1) if decided else None,
                "avg_approval_hours": avg_hours,
                "avg_approval_note": (
                    None if avg_hours is not None
                    else "No submit/approve events on record for this selection"
                ),
                "avg_discount_percent": round(sum(discounts) / len(discounts), 2) if discounts else 0.0,
                "total_discount_amount": round(total_discount_amount, 2),
                "high_discount_count": len(high_disc_lines),
                "critical_discount_count": len(critical_disc_lines),
                "top_product": best_selling[0]["name"] if best_selling else None,
            },
            "data_quality": {
                "unlinked_lines": unlinked.get("lines", 0),
                "unlinked_revenue": round(unlinked.get("revenue", 0.0), 2),
                "note": (
                    "Lines with no catalog link are excluded from product rankings. "
                    "Run scripts/backfill_line_links.py to reduce this count."
                ) if unlinked.get("lines") else None,
            },
            "status_breakdown": [
                {"status": k, "count": v} for k, v in sorted(statuses.items(), key=lambda kv: -kv[1])
            ],
            "funnel_stages": funnel_stages,
            "tier_breakdown": tier_breakdown,
            "top_customers": top_customers,
            "monthly_trend": monthly_trend,
            "best_selling": best_selling,
            "most_discounted": most_discounted,
            "reps": reps,
            "categories": [{"id": k, "name": v} for k, v in sorted(categories.items())],
        }
    finally:
        session.close()


# ─────────────────────────────────────────────────────────────────────
# Exports
# ─────────────────────────────────────────────────────────────────────
def _stamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d-%H%M")


@router.get("/analytics/export.xlsx")
def export_xlsx(
    period: str = "month",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rep: Optional[str] = None,
    approval_status: Optional[str] = None,
    category_id: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(
        RoleChecker([RoleEnum.admin, RoleEnum.manager, RoleEnum.finance])
    ),
):
    """Multi-sheet workbook: KPIs, best selling, most discounted, rep performance."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    data = sales_analytics(
        period, date_from, date_to, rep, approval_status, category_id, product_id, current_user
    )

    wb = Workbook()
    header_fill = PatternFill("solid", fgColor="714B67")   # Odoo plum
    header_font = Font(bold=True, color="FFFFFF")

    def sheet(title, columns, rows):
        ws = wb.create_sheet(title)
        ws.append(columns)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for r in rows:
            ws.append(r)
        for i, col in enumerate(columns, start=1):
            width = max(len(str(col)) + 2, *(len(str(r[i - 1])) + 2 for r in rows)) if rows else len(col) + 2
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width, 52)
        ws.freeze_panes = "A2"
        return ws

    wb.remove(wb.active)

    k = data["kpis"]
    sheet("Summary", ["Metric", "Value"], [
        ["Period", data["filters"]["period"]],
        ["Rep filter", data["filters"]["rep"] or "All"],
        ["Approval status", data["filters"]["approval_status"] or "All"],
        ["Category", data["filters"]["category_id"] or "All"],
        ["Quotes created", k["quotes_created"]],
        ["Total value (INR)", k["total_value"]],
        ["Average deal size (INR)", k["average_deal_size"]],
        ["Won", k["won_count"]],
        ["Win rate %", k["win_rate"] if k["win_rate"] is not None else "n/a"],
        ["Avg approval hours", k["avg_approval_hours"] if k["avg_approval_hours"] is not None else k["avg_approval_note"]],
        ["Avg discount %", k["avg_discount_percent"]],
        ["Top product", k["top_product"] or "n/a"],
        ["Generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ])

    sheet("Best Selling", ["Product", "Category", "Units", "Revenue (INR)", "Avg Discount %"],
          [[p["name"], p["category"] or "-", p["units"], p["revenue"], p["avg_discount"]] for p in data["best_selling"]])

    sheet("Most Discounted", ["Product", "Category", "Avg Discount %", "Units", "Revenue (INR)"],
          [[p["name"], p["category"] or "-", p["avg_discount"], p["units"], p["revenue"]] for p in data["most_discounted"]])

    sheet("Rep Performance", ["Rep", "Quotes", "Value (INR)", "Won", "Win Rate %"],
          [[r["rep"], r["quotes"], r["value"], r["won"], r["win_rate"]] for r in data["reps"]])

    sheet("Customer Tiers", ["Customer Tier", "Quotes", "Pipeline Value (INR)", "Won Value (INR)", "Won Count", "Win Rate %", "Avg Discount %"],
          [[t["tier"], t["quotes"], t["value"], t["won_value"], t["won_count"], t["win_rate"], t["avg_discount"]] for t in data.get("tier_breakdown", [])])

    sheet("Top Accounts", ["Account Name", "Tier", "Quotes", "Pipeline Value (INR)", "Won Value (INR)"],
          [[c["name"], c["tier"], c["quotes"], c["value"], c["won_value"]] for c in data.get("top_customers", [])])

    sheet("Funnel Breakdown", ["Stage", "Count", "Value (INR)", "Pipeline Share %"],
          [[s["stage"], s["count"], s["value"], s["share_pct"]] for s in data.get("funnel_stages", [])])

    sheet("Status Breakdown", ["Status", "Count"],
          [[s["status"], s["count"]] for s in data["status_breakdown"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="dealflow360-report-{_stamp()}.xlsx"'},
    )


@router.get("/analytics/export.pdf")
def export_pdf(
    period: str = "month",
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    rep: Optional[str] = None,
    approval_status: Optional[str] = None,
    category_id: Optional[str] = None,
    product_id: Optional[str] = None,
    current_user: dict = Depends(
        RoleChecker([RoleEnum.admin, RoleEnum.manager, RoleEnum.finance])
    ),
):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    data = sales_analytics(
        period, date_from, date_to, rep, approval_status, category_id, product_id, current_user
    )

    PLUM = colors.HexColor("#714B67")
    INK = colors.HexColor("#212529")
    styles = getSampleStyleSheet()
    title = ParagraphStyle("T", parent=styles["Normal"], fontName="Helvetica-Bold",
                           fontSize=18, leading=22, textColor=PLUM)
    sub = ParagraphStyle("S", parent=styles["Normal"], fontName="Helvetica",
                         fontSize=9, leading=12, textColor=colors.HexColor("#6C757D"))
    h2 = ParagraphStyle("H2", parent=styles["Normal"], fontName="Helvetica-Bold",
                        fontSize=12, leading=15, textColor=INK, spaceBefore=14, spaceAfter=6)

    def table(columns, rows, widths):
        t = Table([columns] + rows, colWidths=widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), PLUM),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#DEE2E6")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8F4F7")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5),
            ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ]))
        return t

    f, k = data["filters"], data["kpis"]
    story = [
        Paragraph("DealFlow360 — Sales Report", title),
        Paragraph(
            f"Period: {f['period']} &nbsp;|&nbsp; Rep: {f['rep'] or 'All'} &nbsp;|&nbsp; "
            f"Approval: {f['approval_status'] or 'All'} &nbsp;|&nbsp; "
            f"Category: {f['category_id'] or 'All'} &nbsp;|&nbsp; "
            f"Generated {datetime.now(timezone.utc).strftime('%d %b %Y %H:%M UTC')}",
            sub,
        ),
        Spacer(1, 14),
        table(
            ["Quotes", "Total Value (INR)", "Avg Deal (INR)", "Won", "Win Rate", "Avg Approval", "Avg Discount"],
            [[
                k["quotes_created"],
                f"{k['total_value']:,.0f}",
                f"{k['average_deal_size']:,.0f}",
                k["won_count"],
                f"{k['win_rate']}%" if k["win_rate"] is not None else "n/a",
                f"{k['avg_approval_hours']} h" if k["avg_approval_hours"] is not None else "no data",
                f"{k['avg_discount_percent']}%",
            ]],
            [50, 92, 82, 40, 55, 68, 62],
        ),
    ]

    if data["best_selling"]:
        story += [
            Paragraph("Best Selling Products", h2),
            table(["Product", "Category", "Units", "Revenue (INR)", "Avg Disc %"],
                  [[p["name"][:46], (p["category"] or "-")[:18], p["units"],
                    f"{p['revenue']:,.0f}", p["avg_discount"]] for p in data["best_selling"][:8]],
                  [200, 90, 45, 90, 60]),
        ]

    if data["most_discounted"]:
        story += [
            Paragraph("Most Discounted Products", h2),
            table(["Product", "Category", "Avg Disc %", "Units"],
                  [[p["name"][:46], (p["category"] or "-")[:18], p["avg_discount"], p["units"]]
                   for p in data["most_discounted"][:8]],
                  [220, 100, 80, 60]),
        ]

    if data["reps"]:
        story += [
            Paragraph("Rep Performance", h2),
            table(["Rep", "Quotes", "Value (INR)", "Won", "Win Rate %"],
                  [[r["rep"][:36], r["quotes"], f"{r['value']:,.0f}", r["won"], r["win_rate"]]
                   for r in data["reps"][:10]],
                  [190, 55, 100, 50, 70]),
        ]

    buf = io.BytesIO()
    SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=32, rightMargin=32, topMargin=36, bottomMargin=36,
        title="DealFlow360 Sales Report",
    ).build(story)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="dealflow360-report-{_stamp()}.pdf"'},
    )
